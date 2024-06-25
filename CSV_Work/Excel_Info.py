""" A script to report information about a specified Excel file. """

import openpyxl           # The required import library
import sys                # The library for system functions
import os                 # A library of system functions
from pathlib import Path  # The library for file/path operations

# Obtain the current working directory, which will be the assumed default
# directory if the user specifies only a file name.

cwd = Path.cwd()

# Acquire the Excel file name (or full pathname) from the command line.
if len( sys.argv ) > 1:
    excel_file = " ".join( sys.argv[1:] )
else:
    print( "\n You must specify a file or pathname on the command line.\n" )
    sys.exit()

# Check if the command line argument is a filename or a pathname.  For filenames,
# prepend with the current working directory to build the path.

if not( os.path.isabs( excel_file ) ):
    excel_file = os.path.join( cwd, excel_file )  #combine the cwd and filename

print( f"\n The specified Excel file pathname is: {excel_file}." )


# Open the Excel file and report its characteristics.
efile  = openpyxl.load_workbook( excel_file )

sheets = efile.sheetnames
num_sheets = len( sheets )
print( f" There are {num_sheets} in this Excel file. The sheet names are:" )
print( sheets )

for i in range( num_sheets ):
    sheet = efile[sheets[i]]
    print( f" For {sheet} there are {sheet.max_row} rows and {sheet.max_column} columns." )











